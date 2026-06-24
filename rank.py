#!/usr/bin/env python3
"""Phase 8 — THE RANKING STEP. Produces submission.csv. CPU-only, no network, <5 min.

For the official candidates.jsonl every id is in the precomputed feature store, so this is a
parquet load + cheap numpy + one LightGBM predict + a top-100 reasoning pass. Candidates not
in the store (sandbox uploads) get features built live. Deterministic: sort by final score
desc, tie-break candidate_id ascending.

Usage: python rank.py --candidates ./candidates.jsonl --out ./submission.csv
"""
from __future__ import annotations

import argparse
import csv
import time
import tracemalloc

import pandas as pd

from src.features.registry import get_reference_now  # noqa: F401 (warms config)
from src.io.artifacts import load_features
from src.io.config import load_config, resolve_path
from src.io.loaders import iter_candidates
from src.scoring import pipeline, ranker
from src.scoring.reasoning import reason_for


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--candidates", required=True)
    ap.add_argument("--out", default="submission.csv")
    ap.add_argument("--top", type=int, default=100)
    args = ap.parse_args()

    t0 = time.time()
    tracemalloc.start()
    cfg = load_config()

    # target ids = ids present in the candidates file (preserve existence guarantee)
    target_ids = [c.candidate_id for c in iter_candidates(args.candidates)]
    target_set = set(target_ids)

    store = load_features(resolve_path(cfg["artifacts"]["features"]))
    cached = store[store.index.isin(target_set)]
    missing = target_set - set(cached.index)

    if missing:
        # sandbox / uncached path: build features live (loads embedder)
        from src.live_features import build_live_features
        miss_cands = [c for c in iter_candidates(args.candidates) if c.candidate_id in missing]
        live = build_live_features(miss_cands, cfg)
        df = pd.concat([cached, live])
    else:
        df = cached
    print(f"[{time.time()-t0:.1f}s] features ready for {len(df):,} candidates "
          f"({len(missing)} live-built)")

    model, cols = (None, None)
    if cfg["ranker"].get("enabled", True):
        model, cols = ranker.load(resolve_path(cfg["artifacts"]["ranker"]))
    scored = pipeline.score(df, cfg, model=model, cols=cols)
    top = pipeline.rank_top(scored, top_n=args.top)
    print(f"[{time.time()-t0:.1f}s] scored + ranked  (model={'lgbm+rubric' if model else 'rubric-only'})")

    # fetch Candidate objects for the top-N (for grounded reasoning)
    top_ids = set(top["candidate_id"])
    objs = {c.candidate_id: c for c in iter_candidates(args.candidates)
            if c.candidate_id in top_ids}

    rows = []
    for _, r in top.iterrows():
        cid = r["candidate_id"]
        feat = r.to_dict()
        reasoning = reason_for(objs[cid], feat, int(r["rank"]))
        rows.append((cid, int(r["rank"]), round(float(r["final"]), 6), reasoning))

    with open(args.out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["candidate_id", "rank", "score", "reasoning"])
        w.writerows(rows)

    # XLSX deliverable (same content, formatted) alongside the validator-facing CSV.
    xlsx_path = args.out[:-4] + ".xlsx" if args.out.endswith(".csv") else args.out + ".xlsx"
    write_xlsx(xlsx_path, rows)

    cur, peak = tracemalloc.get_traced_memory()
    print(f"[{time.time()-t0:.1f}s] wrote {args.out} + {xlsx_path}  ({len(rows)} rows)")
    print(f"wall={time.time()-t0:.1f}s  peak_python_mem={peak/1e6:.0f}MB")


def write_xlsx(path: str, rows: list[tuple]) -> None:
    """Ranked top-N as a clean .xlsx: bold header, sized columns, wrapped reasoning."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"
    headers = ["candidate_id", "rank", "score", "reasoning"]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(vertical="center")
    for cid, rank, score, reasoning in rows:
        ws.append([cid, rank, score, reasoning])
    widths = {"A": 16, "B": 6, "C": 11, "D": 110}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[1].alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    main()
